import numpy as np
import openpyxl as excel
from openpyxl.utils import get_column_letter
import FinanceDataReader as fdr
from tqdm import tqdm

test = np.load('test_npy.npy')
dict = {}

# ------------------------------- Common function ---------------------------------------------------
Today = 20230315 # insert Today YYYYMMDD
Y_value = int(str(Today)[:4])
Q_value = int(str(Today)[4:])
year_range = 10
quarter = ['0331', '0630', '0930', '1231']
start_idx = 0 if Q_value <= int(quarter[0]) else 1 if Q_value <= int(quarter[1]) else 2 if Q_value <= int(quarter[2]) else 3
start_idx_copy = start_idx
Cal_x = []
for year in range(Y_value, Y_value - year_range-1, -1):
    while not start_idx < 0:
        temp_q = str(year) + quarter[start_idx]
        Cal_x.append(temp_q)
        start_idx -= 1
    start_idx = 3

for value in test:
    if value[0] in dict.keys():
        dict[value[0]].append([int(value[1]), float(value[2])])
    else:
        dict[value[0]] = [[int(value[1]), float(value[2])]]
print(dict.keys())
test_set = list(dict.keys())
# ------------------------------- To make allocation list ------------------------------------
ap_set = [] # Allocation list
for Comp in test_set:
    temp = list(dict[Comp])
    asc_temp = sorted(temp, key=lambda temp: temp[0])
    cal_data = []
    for date in Cal_x:
        account = 0
        temp_date = 0
        if asc_temp != []:
            while not int(date) > asc_temp[-1][0]:
                if temp_date != asc_temp[-1][0]:
                    account += asc_temp[-1][1]
                    temp_date = asc_temp[-1][0]
                asc_temp.pop()
                if asc_temp == []:
                    account = 0
                    break
        cal_data.append(account)
    ap_set.append(cal_data)

# ------------------------------- To make allocation list ------------------------------------
stock_set = [] # stock price list
start_day = str(int(Today) - year_range*10000)
stock = fdr.StockListing('KOSPI')
Find_id = stock['Name']
cmp_code = stock['Code']

marcap_list = []
stock_account = []
non_value = len(Cal_x)*[-1]
process_list = tqdm(test_set)

for ccc in process_list:
    temp = list(Find_id.index[np.where(Find_id == ccc)])
    if temp != []:
        on_value = []
        iid = int(temp[0])
        cmp = cmp_code[iid]
        # - Close account track
        stock_data = fdr.DataReader(cmp, start_day, str(Today))
        Close_slice = stock_data['Close']
        temp_date = []
        for date in Close_slice.index.values:
            temp_date.append(str(date).replace('-', '')[:8])
        d_c_value = np.array([temp_date, Close_slice.values]).T.tolist()
        # Averaging by Quarter
        for date in Cal_x:
            ac_plus = [0]
            if d_c_value != []:
                while not int(date)-300 > int(d_c_value[-1][0]):
                    ac_plus.append(int(d_c_value[-1][1]))
                    d_c_value.pop()
                    if d_c_value == []:
                        break
            on_value.append(np.round(np.sum(ac_plus) / len(ac_plus), 2))
        marcap_list.append(int(stock['Marcap'][temp]))
        stock_account.append(int(stock['Stocks'][temp]))
        stock_set.append(on_value)
    else:
        marcap_list.append(-1)
        stock_account.append(-1)
        stock_set.append(non_value)
    process_list.set_description(f'Processing {ccc}')

# ----------------------------- Make excel -----------------------------
wb = excel.Workbook()
wb.active.title = 'Allocation'
wb.create_sheet('Stock_price', 1)
wb.create_sheet('All_ratio', 2)
w1 = wb['Allocation']
w2 = wb['Stock_price']
w3 = wb['All_ratio']

Q = ['Q1', 'Q2', 'Q3', 'Q4']
x_axis = ['시가 총액', '주식량 (주)']
for Q_date in Cal_x:
    for iid,  match in enumerate(quarter):
        if Q_date[4:] == match:
            temp = 'FY' + Q_date[2:4] + Q[iid]
            x_axis.append(temp)

w1.cell(1, 1).value = '배당 기록 (원)'
w2.cell(1, 1).value = '수정 주가 기록 (주)'
w3.cell(1, 1).value = '주당배당률'

for idx, day in enumerate(x_axis):
    w1.cell(1, idx + 2).value = day
    w2.cell(1, idx + 2).value = day
    if idx < 2:
        w3.cell(1, idx + 2).value = day

full_length = len(Cal_x)
c_row = 0
for idx, company in enumerate(test_set):
    if marcap_list[idx] != -1:
        w1.cell(2 + c_row, 1).value = company
        w2.cell(2 + c_row, 1).value = company
        w3.cell(2 + c_row, 1).value = company

        w1.cell(2 + c_row, 2).value = marcap_list[idx]
        w2.cell(2 + c_row, 2).value = marcap_list[idx]
        w3.cell(2 + c_row, 2).value = marcap_list[idx]

        w1.cell(2 + c_row, 3).value = stock_account[idx]
        w2.cell(2 + c_row, 3).value = stock_account[idx]
        w3.cell(2 + c_row, 3).value = stock_account[idx]

        Year_allocate, Year_stock, Year_count = [], [], 0
        start_idx = start_idx_copy
        for i in range(full_length):
            w1.cell(2 + c_row, 4 + i).value = ap_set[idx][i]
            w2.cell(2 + c_row, 4 + i).value = stock_set[idx][i]
            Year_allocate.append(ap_set[idx][i])
            Year_stock.append(stock_set[idx][i])
            if start_idx == 0:
                w3.cell(1, Year_count + 4).value = int(Cal_x[i][:4])
                if np.average(Year_stock) != 0:
                    Year_ratio = np.sum(Year_allocate) / np.average(Year_stock) * 100
                else:
                    Year_ratio = 0.00
                w3.cell(2 + c_row, 4 + Year_count).value = str(round(Year_ratio, 2)) + "%"
                Year_allocate, Year_stock = [], []
                start_idx = 4
                Year_count += 1
            start_idx -= 1
        c_row += 1

for col in range(1, 4):
    w1.column_dimensions[get_column_letter(col)].width = 18
    w2.column_dimensions[get_column_letter(col)].width = 18
    w3.column_dimensions[get_column_letter(col)].width = 18

wb.save("Plan.xlsx")