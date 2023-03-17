import numpy as np
import openpyxl as excel

test = np.load('meta_data.npy')

dict = {}
np_set = []

Today = 20230315 # insert Today YYYYMMDD
Y_value = int(str(Today)[:4])
Q_value = int(str(Today)[4:])
year_range = 10
quarter = ['0331', '0630', '0930', '1231']
start_idx = 0 if Q_value <= int(quarter[0]) else 1 if Q_value <= int(quarter[1]) else 2 if Q_value <= int(quarter[2]) else 3
Cal_x = []
for year in range(Y_value, Y_value - year_range-1, -1):
    while not start_idx < 0:
        temp_q = str(year) + quarter[start_idx]
        Cal_x.append(temp_q)
        start_idx -= 1
    start_idx = 3
print(Cal_x)

for value in test:
    if value[0] in dict.keys():
        dict[value[0]].append([int(value[1]), float(value[2])])
    else:
        dict[value[0]] = [[int(value[1]), float(value[2])]]
print(dict.keys())

for Comp in dict.keys():
    temp = list(dict[Comp])
    asc_temp = sorted(temp, key=lambda temp: temp[0])
    cal_data = []
    for date in Cal_x:
        account = 0
        temp_date = 0
        if asc_temp != []:
            while not int(date) > asc_temp[-1][0]:
                if temp_date != asc_temp[-1][0]:
                    account += asc_temp[-1][1]
                    temp_date = asc_temp[-1][0]
                asc_temp.pop()
                if asc_temp == []:
                    account = 0
                    break
        cal_data.append(account)
    np_set.append(cal_data)

wb = excel.Workbook()
wb.active.title = 'label'
w1 = wb['label']

Q = ['Q1', 'Q2', 'Q3', 'Q4']
x_axis = []
for Q_date in Cal_x:
    for iid,  match in enumerate(quarter):
        if Q_date[4:] == match:
            temp = 'FY' + Q_date[2:4] + Q[iid]
            x_axis.append(temp)

for idx, day in enumerate(x_axis):
    w1.cell(1, idx + 2).value = day

for idx, company in enumerate(dict.keys()):
    w1.cell(2 + idx, 1).value = company
    for id, money in enumerate(np_set[idx]):
        w1.cell(2 + idx, 2 + id).value = money

wb.save("Stock_result2.xlsx")
