import numpy as np
import openpyxl as excel
import FinanceDataReader as fdr

test = np.load('test_npy.npy')
dict = {}

Today = 20230315 # insert Today YYYYMMDD
Y_value = int(str(Today)[:4])
Q_value = int(str(Today)[4:])
year_range = 10
start_day = str(int(Today) - year_range*10000)
quarter = ['0331', '0630', '0930', '1231']
start_idx = 0 if Q_value <= int(quarter[0]) else 1 if Q_value <= int(quarter[1]) else 2 if Q_value <= int(quarter[2]) else 3
Cal_x = []
for year in range(Y_value, Y_value - year_range-1, -1):
    while not start_idx < 0:
        temp_q = str(year) + quarter[start_idx]
        Cal_x.append(temp_q)
        start_idx -= 1
    start_idx = 3
print(Cal_x)
for value in test:
    if value[0] in dict.keys():
        dict[value[0]].append([int(value[1]), float(value[2])])
    else:
        dict[value[0]] = [[int(value[1]), float(value[2])]]


c_list = list(dict.keys())
stock = fdr.StockListing('KOSPI')
Find_id = stock['Name']
cmp_code = stock['Code']

# test by ss
iid = int(Find_id.index[np.where(Find_id == '삼성전자')[0][0]])
cmp = cmp_code[iid]
data1 = fdr.DataReader(cmp, start_day, str(Today))
print(data1.columns)
print(data1)
Close_slice = data1['Close']
print(Close_slice)



# stock_set = []
# non_value = len(Cal_x)*[-1]
#
# for ccc in c_list:
#     temp = list(Find_id.index[np.where(Find_id == ccc)])
#     if temp != []:
#         print(ccc)
#         on_value = []
#         iid = int(temp[0])
#         cmp = cmp_code[iid]
#         # - Close account track
#         stock_data = fdr.DataReader(cmp, start_day, str(Today))
#         Close_slice = stock_data['Close']
#         temp_date = []
#         for date in Close_slice.index.values:
#             temp_date.append(str(date).replace('-', '')[:8])
#         d_c_value = np.array([temp_date, Close_slice.values]).T.tolist()
#         # Averaging by Quarter
#         for date in Cal_x:
#             ac_plus = [0]
#             if d_c_value != []:
#                 while not int(date)-300 > int(d_c_value[-1][0]):
#                     ac_plus.append(int(d_c_value[-1][1]))
#                     d_c_value.pop()
#                     if d_c_value == []:
#                         break
#             on_value.append(np.round(np.sum(ac_plus) / len(ac_plus), 2))
#         stock_set.append(on_value)
#     else:
#         stock_set.append(non_value)
#
# wb = excel.Workbook()
# wb.active.title = 'label'
# w1 = wb['label']
#
# Q = ['Q1', 'Q2', 'Q3', 'Q4']
# x_axis = []
# for Q_date in Cal_x:
#     for iid,  match in enumerate(quarter):
#         if Q_date[4:] == match:
#             temp = 'FY' + Q_date[2:4] + Q[iid]
#             x_axis.append(temp)
#
# for idx, day in enumerate(x_axis):
#     w1.cell(1, idx + 2).value = day
#
# for idx, company in enumerate(dict.keys()):
#     w1.cell(2 + idx, 1).value = company
#     for id, money in enumerate(stock_set[idx]):
#         w1.cell(2 + idx, 2 + id).value = money
#
# wb.save("stock.xlsx")